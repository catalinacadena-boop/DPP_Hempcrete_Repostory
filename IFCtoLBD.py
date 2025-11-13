#Made by P. Pauwels (https://github.com/pipauwel/IFCtoLBD)
from datetime import datetime
import ifcopenshell as ios
from ifcopenshell.util.element import get_psets
import Namespace
import os
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import tkinter as tk
from tkinter import filedialog, simpledialog, messagebox 

global baseURI 
inputFile = ""
targetFile = ""

includeBuildingElements = True
includeBuildingProperties = True
includeQuantities = True
includeGeometry = False

# Global list to store Excel data
excel_data = []

# Global dictionary to store element filters
element_filters = {
    'sites': True,
    'buildings': True,
    'storeys': True,
    'spaces': True,
    'elements': True,
    'interfaces': True,
    'zones': True
}

def select_element_types(model):
    """Show dialog to select which element types to process"""
    global element_filters
    
    # Count elements in model
    counts = {
        'Sites': len(list(model.by_type("IfcSite"))),
        'Buildings': len(list(model.by_type("IfcBuilding"))),
        'Storeys': len(list(model.by_type("IfcBuildingStorey"))),
        'Spaces': len(list(model.by_type("IfcSpace"))),
        'Elements': len(list(model.by_type("IfcElement"))),
        'Interfaces': len(list(model.by_type("IfcRelSpaceBoundary"))),
        'Zones': len(list(model.by_type("IfcZone")))
    }
    
    # Create selection dialog
    dialog = tk.Toplevel()
    dialog.title("Select Element Types to Process")
    dialog.geometry("400x350")
    dialog.attributes('-topmost', True)
    
    # Center the dialog
    dialog.update_idletasks()
    x = (dialog.winfo_screenwidth() // 2) - (400 // 2)
    y = (dialog.winfo_screenheight() // 2) - (350 // 2)
    dialog.geometry(f"400x350+{x}+{y}")
    
    tk.Label(dialog, text="Select which element types to include:", 
             font=('Arial', 11, 'bold'), pady=10).pack()
    
    tk.Label(dialog, text="(Found in IFC file)", 
             font=('Arial', 9, 'italic')).pack()
    
    # Create checkboxes
    vars_dict = {}
    frame = tk.Frame(dialog, pady=10)
    frame.pack()
    
    mapping = {
        'Sites': 'sites',
        'Buildings': 'buildings', 
        'Storeys': 'storeys',
        'Spaces': 'spaces',
        'Elements': 'elements',
        'Interfaces': 'interfaces',
        'Zones': 'zones'
    }
    
    for display_name, key in mapping.items():
        var = tk.BooleanVar(value=True)
        vars_dict[key] = var
        count = counts[display_name]
        cb = tk.Checkbutton(frame, 
                           text=f"{display_name} ({count} found)",
                           variable=var,
                           font=('Arial', 10))
        cb.pack(anchor='w', pady=3, padx=20)
    
    # Buttons
    button_frame = tk.Frame(dialog)
    button_frame.pack(pady=20)
    
    result = {'confirmed': False}
    
    def select_all():
        for var in vars_dict.values():
            var.set(True)
    
    def deselect_all():
        for var in vars_dict.values():
            var.set(False)
    
    def confirm():
        result['confirmed'] = True
        for key, var in vars_dict.items():
            element_filters[key] = var.get()
        dialog.destroy()
    
    def cancel():
        dialog.destroy()
    
    tk.Button(button_frame, text="Select All", command=select_all, width=12).pack(side='left', padx=5)
    tk.Button(button_frame, text="Deselect All", command=deselect_all, width=12).pack(side='left', padx=5)
    tk.Button(button_frame, text="OK", command=confirm, width=12, bg='#4CAF50', fg='white').pack(side='left', padx=5)
    tk.Button(button_frame, text="Cancel", command=cancel, width=12).pack(side='left', padx=5)
    
    dialog.protocol("WM_DELETE_WINDOW", cancel)
    dialog.wait_window()
    
    return result['confirmed']

def convertIFCSPFtoTTL(inputFile, outputFile):
    inputFile = inputFile
    outputFile = outputFile
    now = datetime.now()
    current_time = now.strftime('%Y%m%d_%H%M%S')
	
    DEFAULT_PATH = "http://linkedbuildingdata.net/ifc/resources" + current_time + "/"
    global baseURI, excel_data
    baseURI = DEFAULT_PATH
    excel_data = []  # Reset for each conversion

	#Enter the name of the ifc file behind pad between the " ". Just like the example. - reading file - IfcSpfReader
    pad = "Project1.ifc"
    model = ios.open(inputFile)
    
    # Show element type selection dialog
    if not select_element_types(model):
        print("Element selection cancelled. Using all element types.")

    f = open(outputFile, "w", encoding='utf-8')
    writeTTLFileContent(model,f)    
    f.close()
    
    # Generate Excel file
    excelFile = outputFile.replace('.ttl', '.xlsx')
    generate_excel_output(excelFile)
    
def writeTTLFileContent(model, file):
    file.write(writeTTLHeader())
    file.write(writeLBDinstances(model,file))

def print_properties(properties, output, element_id="", element_type="", element_name=""):    
    global excel_data
    for name, value in properties.items():   
        if name == "id":
            continue     
        name = cleanString(name)
        output += ";\n"
        
        # Determine data type and unit
        data_type = ""
        unit = ""
        
        if isinstance(value, bool):
            output += "\tprops:"+name+" \""+ str(value) +"\"^^xsd:boolean "
            data_type = "boolean"
        elif isinstance(value, int):          
            output += "\tprops:"+name+" \""+ str(value) +"\"^^xsd:int "
            data_type = "integer"
            unit = extract_unit_from_name(name)
        elif isinstance(value, float):    
            output += "\tprops:"+name+" \""+ str(value) +"\"^^xsd:double "
            data_type = "double"
            unit = extract_unit_from_name(name)
        else:
            # Escape newlines and quotes in strings
            value_str = str(value).replace('\n', ', ').replace('\r', '').replace('"', '\\"')
            output += "\tprops:"+name+" \""+ value_str +"\"^^xsd:string "
            data_type = "string"
        
        # Add to Excel data
        excel_data.append({
            'Element_ID': element_id,
            'Element_Type': element_type,
            'Element_Name': element_name,
            'Parameter': name,
            'Value': str(value),
            'Data_Type': data_type,
            'Unit': unit
        })
    return output

""" def print_quantities(quantities, output):    
    for name, value in quantities.items():   
        if name == "id":
            continue     
        name = cleanString(name)
        output += ";\n"
        if isinstance(value, int):          
            output += "\tprops:"+name+" \""+ str(value) +"\"^^xsd:int "
        elif isinstance(value, float):    
            output += "\tprops:"+name+" \""+ str(value) +"\"^^xsd:double "
        else:           
            output += "\tprops:"+name+" \""+ str(value) +"\"^^xsd:string "
    return output """

def cleanString(name):
    name = ''.join(x for x in name.title() if not x.isspace())
    name = name.replace('\\', '')
    name = name.replace('/', '')
    return name

def extract_unit_from_name(param_name):
    """Extract unit from parameter name if present"""
    # Use regex patterns to match units more precisely
    # Check for specific patterns with word boundaries or underscores
    import re
    
    # Order matters - check more specific patterns first
    unit_patterns = [
        (r'_Kgco₂?Eq', 'kgCO₂eq'),
        (r'_Kgso₂?Eq', 'kgSO₂eq'),
        (r'_Mjkg', 'MJ/kg'),
        (r'_Mpa', 'MPa'),
        (r'_Eur', 'EUR'),
        (r'_Years', 'years'),
        (r'_Ctuh', 'CTUh'),
        (r'_Mm(?:$|_)', 'mm'),  # Match Mm at end or before underscore
        (r'_Cm(?:$|_)', 'cm'),
        (r'_Km(?:$|_)', 'km'),
        (r'_Kg(?:$|_)', 'kg'),
        (r'_M(?:$|_)', 'm'),    # Match single M at end or before underscore
        (r'_M2', 'm²'),
        (r'_M3', 'm³'),
        (r'_M²', 'm²'),
        (r'_M³', 'm³'),
    ]
    
    for pattern, unit in unit_patterns:
        if re.search(pattern, param_name, re.IGNORECASE):
            return unit
    
    return ""

def generate_excel_output(excel_file):
    """Generate Excel file with all parameters, values and units"""
    global excel_data
    
    if not excel_data:
        print("No data to export to Excel")
        return
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IFC Parameters"
    
    # Define headers
    headers = ['Element ID', 'Element Type', 'Element Name', 'Parameter', 'Value', 'Data Type', 'Unit']
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Write data
    for row_num, data_row in enumerate(excel_data, 2):
        ws.cell(row=row_num, column=1, value=data_row['Element_ID'])
        ws.cell(row=row_num, column=2, value=data_row['Element_Type'])
        ws.cell(row=row_num, column=3, value=data_row['Element_Name'])
        ws.cell(row=row_num, column=4, value=data_row['Parameter'])
        ws.cell(row=row_num, column=5, value=data_row['Value'])
        ws.cell(row=row_num, column=6, value=data_row['Data_Type'])
        ws.cell(row=row_num, column=7, value=data_row['Unit'])
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Set default view to start at cell A1
    ws.sheet_view.showGridLines = True
    ws.sheet_view.showRowColHeaders = True
    ws.sheet_view.tabSelected = True
    
    # Set print settings
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    
    # Save workbook
    wb.save(excel_file)
    print(f"Excel file generated: {excel_file}")
    print(f"Total rows exported: {len(excel_data)}")



def writeTTLHeader():
    s = "# baseURI: " + baseURI + "\n"
    s+= "@prefix inst: <" + baseURI + "> .\n"
    s+= "@prefix rdf:  <" + Namespace.RDF + "> .\n"
    s+= "@prefix rdfs:  <" + Namespace.RDFS + "> .\n"
    s+= "@prefix xsd:  <" + Namespace.XSD + "> .\n"
    s+= "@prefix bot:  <" + Namespace.BOT + "> .\n"
    s+= "@prefix beo:  <" + Namespace.BEO + "> .\n"
    s+= "@prefix mep:  <" + Namespace.MEP + "> .\n"
    s+= "@prefix geom:  <" + Namespace.GEOM + "> .\n"
    s+= "@prefix props:  <" + Namespace.PROPS + "> .\n\n"

    s+= "inst: rdf:type <http://www.w3.org/2002/07/owl#Ontology> .\n\n"

    return s

def writeLBDinstances(model, f):
    global element_filters
    output = ""
    
    if element_filters.get('sites', True):
        output += writeSites(model,f)
        print("  ✓ Processed Sites")
    
    if element_filters.get('buildings', True):
        output += writeBuildings(model,f)
        print("  ✓ Processed Buildings")
    
    if element_filters.get('storeys', True):
        output += writeStoreys(model,f)
        print("  ✓ Processed Storeys")
    
    if element_filters.get('spaces', True):
        output += writeSpaces(model,f)
        print("  ✓ Processed Spaces")
    
    if element_filters.get('elements', True):
        output += writeElements(model,f)
        print("  ✓ Processed Elements")
    
    if element_filters.get('interfaces', True):
        output += writeInterfaces(model,f)
        print("  ✓ Processed Interfaces")
    
    if element_filters.get('zones', True):
        output += writeZones(model,f)
        print("  ✓ Processed Zones")
    
    return output

def writeSites(model,f):
    output = ""
    for s in model.by_type("IfcSite"):                
        site_id = "site_"+str(s.id())
        site_name = s.Name if s.Name else ""
        output += "inst:" + site_id + "\n"
        output += "\ta bot:Site ;" + "\n"
        if(s.Name):
            output += "\trdfs:label \""+s.Name+"\"^^xsd:string ;" + "\n"
        if(s.Description):
            output += "\trdfs:comment \""+s.Description+"\"^^xsd:string ;" + "\n"        
        output += "\tbot:hasGuid \""+ ios.guid.expand(s.GlobalId) +"\"^^xsd:string ;" + "\n" # bot:hasGuid no such property in the bot ontology？
        output += "\tprops:hasCompressedGuid \""+ s.GlobalId +"\"^^xsd:string "
        for reldec in s.IsDecomposedBy:
            if reldec is not None:
                for b in reldec.RelatedObjects:
                    output += ";\n"
                    output += "\tbot:hasBuilding inst:building_"+ str(b.id()) + " "
        if(includeBuildingProperties):
            site_psets = get_psets(s)
            for name, properties in site_psets.items():
                output = print_properties(properties, output, site_id, "Site", site_name)                             
                
        output += ". \n\n"
    return output

def writeBuildings(model,f):
    output = ""
    for b in model.by_type("IfcBuilding"):                
        building_id = "building_"+str(b.id())
        building_name = b.Name if b.Name else ""
        output += "inst:" + building_id + "\n"
        output += "\ta bot:Building ;" + "\n"
        if(b.Name):
            output += "\trdfs:label \""+b.Name+"\"^^xsd:string ;" + "\n"
        if(b.Description):
            output += "\trdfs:comment \""+b.Description+"\"^^xsd:string ;" + "\n"        
        output += "\tbot:hasGuid \""+ ios.guid.expand(b.GlobalId) +"\"^^xsd:string ;" + "\n"
        output += "\tprops:hasCompressedGuid \""+ b.GlobalId +"\"^^xsd:string "
        for reldec in b.IsDecomposedBy:
            if reldec is not None:
                for st in reldec.RelatedObjects:
                    output += ";\n"
                    output += "\tbot:hasStorey inst:storey_"+ str(st.id()) + " "
        if(includeBuildingProperties):
            psets = get_psets(b)
            for name, properties in psets.items():
                output = print_properties(properties, output, building_id, "Building", building_name)                             
                
        output += ". \n\n"
    return output

def writeStoreys(model,f):
    output = ""
    for b in model.by_type("IfcBuildingStorey"):                
        storey_id = "storey_"+str(b.id())
        storey_name = b.Name if b.Name else ""
        output += "inst:" + storey_id + "\n"
        output += "\ta bot:Storey ;" + "\n"
        if(b.Name):
            output += "\trdfs:label \""+b.Name+"\"^^xsd:string ;" + "\n"
        if(b.Description):
            output += "\trdfs:comment \""+b.Description+"\"^^xsd:string ;" + "\n"        
        output += "\tbot:hasGuid \""+ ios.guid.expand(b.GlobalId) +"\"^^xsd:string ;" + "\n"
        output += "\tprops:hasCompressedGuid \""+ b.GlobalId +"\"^^xsd:string "
        for reldec in b.IsDecomposedBy:
            if reldec is not None:
                for st in reldec.RelatedObjects:
                    output += ";\n"
                    output += "\tbot:hasSpace inst:space_"+ str(st.id()) + " "
        for relcontains in b.ContainsElements:
            if relcontains is not None:
                for st in relcontains.RelatedElements:
                    output += ";\n"
                    output += "\tbot:containsElement inst:element_"+ str(st.id()) + " "

        if(includeBuildingProperties):
            psets = get_psets(b)
            for name, properties in psets.items():
                output = print_properties(properties, output, storey_id, "Storey", storey_name)                             
                
        output += ". \n\n"
    return output

def writeSpaces(model,f):
    output = ""
    for b in model.by_type("IfcSpace"):                
        space_id = "space_"+str(b.id())
        space_name = b.Name if b.Name else ""
        output += "inst:" + space_id + "\n"
        output += "\ta bot:Space ;" + "\n"
        if(b.Name):
            output += "\trdfs:label \""+b.Name+"\"^^xsd:string ;" + "\n"
        if(b.Description):
            output += "\trdfs:comment \""+b.Description+"\"^^xsd:string ;" + "\n"        
        output += "\tbot:hasGuid \""+ ios.guid.expand(b.GlobalId) +"\"^^xsd:string ;" + "\n"
        output += "\tprops:hasCompressedGuid \""+ b.GlobalId +"\"^^xsd:string "
        for relbounded in b.BoundedBy:
            if relbounded is not None:
                st = relbounded.RelatedBuildingElement
                output += ";\n"
                output += "\tbot:adjacentElement inst:element_"+ str(st.id()) + " "
        for relcontains in b.ContainsElements:
            if relcontains is not None:
                counter = 0
                for st in relcontains.RelatedElements:
                    if counter == 0:
                        output += ";\n"
                        output += "\tbot:containsElement inst:element_"+ str(st.id()) + " "
                        counter+=1
                    else:
                        output += ", inst:element_"+ str(st.id()) + " "
                        counter+=1

        if(includeBuildingProperties):
            psets = get_psets(b)
            for name, properties in psets.items():
                output = print_properties(properties, output, space_id, "Space", space_name)    

        #if(includeQuantities):
        #    qsets = ios.util.element.get_qsets(b)
        #    for name, quantities in qsets.items():
        #        output = print_quantities(quantities, output)                          
                
        output += ". \n\n"
    return output

def writeZones(model,f):
    output = ""
    for z in model.by_type("ifcZone"):
        zone_id = "zone_" + str(z.id())
        zone_name = z.Name if z.Name else ""
        output += "inst:" + zone_id + "\n"
        output += "\ta bot:Zone ;" + "\n"
        if(z.Name):
            output += "\trdfs:label \""+z.Name+"\"^^xsd:string ;" + "\n"
        if(z.Description):
            output += "\trdfs:comment \""+z.Description+"\"^^xsd:string ;" + "\n"        
        output += "\tprops:hasGuid \""+ ios.guid.expand(z.GlobalId) +"\"^^xsd:string ;" + "\n"
        output += "\tprops:hasCompressedGuid \""+ z.GlobalId +"\"^^xsd:string "
        for reldec in z.IsDecomposedBy:
            if reldec is not None:
                for sp in reldec.RelatedObjects:
                    output += ";\n"
                    output += "\tbot:hasSpace inst:space_"+ str(sp.id()) + " "
        if(includeBuildingProperties):
            psets = get_psets(z)
            for name, properties in psets.items():
                output = print_properties(properties, output, zone_id, "Zone", zone_name)                             
                
        output += ". \n\n"
    return output 

def writeElements(model,f):
    output = ""
    for b in model.by_type("IfcElement"):                
        element_id = "element_"+str(b.id())
        element_name = b.Name if b.Name else ""
        element_type = b.is_a()
        output += "inst:" + element_id + "\n"
        output += "\ta bot:Element ;" + "\n"
        if(b.Name):
            output += "\trdfs:label \""+b.Name+"\"^^xsd:string ;" + "\n"
        if(b.Description):
            output += "\trdfs:comment \""+b.Description+"\"^^xsd:string ;" + "\n"        
        output += "\tbot:hasGuid \""+ ios.guid.expand(b.GlobalId) +"\"^^xsd:string ;" + "\n"
        output += "\tprops:hasCompressedGuid \""+ b.GlobalId +"\"^^xsd:string "

        for relvoids in b.HasOpenings:
            if relvoids is not None:
                st = relvoids.RelatedOpeningElement
                for relfills in st.HasFillings:
                    if relfills is not None:
                        filler = relfills.RelatedBuildingElement
                        output += ";\n"
                        output += "\tbot:hostsElement inst:element_"+ str(filler.id()) + " "

        for relvoids in b.HasOpenings:
            if relvoids is not None:
                st = relvoids.RelatedOpeningElement
                for relfills in st.HasFillings:
                    if relfills is not None:
                        filler = relfills.RelatedBuildingElement
                        output += ";\n"
                        output += "\tbot:hostsElement inst:element_"+ str(filler.id()) + " "

        if(includeBuildingProperties):
            psets = get_psets(b)
            for name, properties in psets.items():
                output = print_properties(properties, output, element_id, element_type, element_name)                             
                
        output += ". \n\n"
    return output

def writeInterfaces(model,f):
    output = ""
    for b in model.by_type("IfcRelSpaceBoundary"):                
        output += "inst:interface_"+str(b.id()) + "\n"
        output += "\ta bot:Interface ;" + "\n"
        if(b.Name):
            output += "\trdfs:label \""+b.Name+"\"^^xsd:string ;" + "\n"
        if(b.Description):
            output += "\trdfs:comment \""+b.Description+"\"^^xsd:string ;" + "\n"        
        output += "\tbot:hasGuid \""+ ios.guid.expand(b.GlobalId) +"\"^^xsd:string ;" + "\n"
        output += "\tprops:hasCompressedGuid \""+ b.GlobalId +"\"^^xsd:string "

        sp = b.RelatingSpace
        el = b.RelatedBuildingElement
        if sp is not None:
            output += ";\n"
            output += "\tbot:interfaceOf inst:space_"+ str(sp.id()) + " "
        if el is not None:
            output += ";\n"
            output += "\tbot:interfaceOf inst:element_"+ str(el.id()) + " "                
                
        output += ". \n\n"
    return output

#Enter the name of the ifc file behind fname between "". Enter the file name in which de document is saved in front. namefile\\
if __name__ == '__main__':
    # Create root window and hide it
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    
    # Select input IFC file
    print("Please select the IFC file to convert...")
    fname = filedialog.askopenfilename(
        title="Select IFC file to convert",
        filetypes=[("IFC files", "*.ifc"), ("All files", "*.*")],
        initialdir="."
    )
    
    if not fname:
        print("No file selected. Exiting...")
        messagebox.showwarning("No File Selected", "No IFC file was selected. Operation cancelled.")
        exit()
    
    print(f"Selected file: {fname}")
    
    # Ask for output file name (without extension)
    default_name = os.path.splitext(os.path.basename(fname))[0]
    output_name = simpledialog.askstring(
        "Output File Name",
        "Enter the name for output files (without extension):",
        parent=root,
        initialvalue=default_name
    )
    
    if not output_name:
        print("No output name provided. Using default...")
        output_name = default_name
    
    # Remove extensions if user added them
    output_name = output_name.replace('.ttl', '').replace('.xlsx', '')
    
    # Ask for output directory
    output_dir = filedialog.askdirectory(
        title="Select folder to save output files",
        initialdir=os.path.dirname(fname)
    )
    
    if not output_dir:
        print("No output directory selected. Using input file directory...")
        output_dir = os.path.dirname(fname)
    
    # Create output file paths
    ofile = os.path.join(output_dir, f"{output_name}.ttl")
    
    print(f"\nConversion Summary:")
    print(f"  Input IFC: {fname}")
    print(f"  Output TTL: {ofile}")
    print(f"  Output Excel: {ofile.replace('.ttl', '.xlsx')}")
    print(f"\nStarting conversion...")
    
    try:
        convertIFCSPFtoTTL(fname, ofile)
        print("\n" + "="*60)
        print("CONVERSION COMPLETE!")
        print("="*60)
        print(f"Files saved successfully:")
        print(f"  - {ofile}")
        print(f"  - {ofile.replace('.ttl', '.xlsx')}")
        print("="*60)
        messagebox.showinfo("Conversion Complete", 
                           f"IFC conversion successful!\n\nFiles saved:\n• {os.path.basename(ofile)}\n• {os.path.basename(ofile.replace('.ttl', '.xlsx'))}\n\nLocation:\n{output_dir}")
    except Exception as e:
        print(f"\nERROR during conversion: {str(e)}")
        import traceback
        traceback.print_exc()
        messagebox.showerror("Conversion Error", f"An error occurred during conversion:\n\n{str(e)}")
    
    root.destroy()

#the converted file is saved in the user-specified location.                                