"""
Excel Dataset Comparison Tool
Compares a base DPP dataset with IFC conversion results
Generates comparison report in Excel and PDF formats
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import tkinter as tk
from tkinter import filedialog, simpledialog, messagebox
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.units import inch
import os
import re

def sanitize_cell_value(value):
    """Sanitize cell value to prevent formula injection and encoding issues"""
    if value is None:
        return ""
    
    # Convert to string
    value_str = str(value).strip()
    
    # Prevent formula injection - if starts with =, +, -, @ add a single quote prefix
    if value_str and value_str[0] in ['=', '+', '-', '@']:
        value_str = "'" + value_str
    
    # Remove any problematic characters that might cause XML issues
    # Keep it simple - only allow basic characters
    value_str = value_str.replace('\x00', '')  # Remove null bytes
    
    return value_str

def normalize_parameter_name(param_name):
    """Normalize parameter names for comparison (remove spaces, underscores, case-insensitive)"""
    if not param_name:
        return ""
    
    # Convert to string and get the last part after colon if exists
    param_str = str(param_name)
    
    # If parameter contains colon (like "DPP_HempBlock_Material:Typ: Dpp_Sad_Fire_Class")
    # Take the part after the last colon
    if ':' in param_str:
        param_str = param_str.split(':')[-1].strip()
    
    # Remove spaces, underscores, convert to lowercase for comparison
    normalized = param_str.replace(' ', '').replace('_', '').lower()
    return normalized

def extract_unit_from_value(value_str):
    """Try to extract unit from value string if embedded"""
    if not isinstance(value_str, str):
        return str(value_str), ""
    
    # Pattern: number followed by unit (e.g., "200 mm", "15.5kgCO₂eq")
    match = re.match(r'^([\d.]+)\s*([a-zA-Z₂₃°µ]+)$', value_str.strip())
    if match:
        return match.group(1), match.group(2)
    return value_str, ""

def normalize_unit(unit):
    """Normalize unit for comparison"""
    if not unit:
        return ""
    
    # Convert to string and strip whitespace
    unit_str = str(unit).strip()
    
    # Treat dashes and empty as "no unit"
    if unit_str in ['—', '-', '–', '―', '']:
        return ""
    
    unit_mapping = {
        'mm': 'mm',
        'millimeter': 'mm',
        'millimeters': 'mm',
        'cm': 'cm',
        'centimeter': 'cm',
        'centimeters': 'cm',
        'm': 'm',
        'meter': 'm',
        'meters': 'm',
        'kg': 'kg',
        'kilogram': 'kg',
        'kilograms': 'kg',
        'kgco2eq': 'kgCO₂eq',
        'kgco₂eq': 'kgCO₂eq',
        'kgso2eq': 'kgSO₂eq',
        'kgso₂eq': 'kgSO₂eq',
        'mjkg': 'MJ/kg',
        'mj/kg': 'MJ/kg',
        'mpa': 'MPa',
        'eur': 'EUR',
        'euro': 'EUR',
        'euros': 'EUR',
        'years': 'years',
        'year': 'years',
        'km': 'km',
        'kilometer': 'km',
        'kilometers': 'km',
        'ctuh': 'CTUh',
        'm²': 'm²',
        'm2': 'm²',
        'm³': 'm³',
        'm3': 'm³',
        'no': 'No.',
        'no.': 'No.',
        'id': 'ID',
        'guid': 'GUID'
    }
    
    normalized = unit_str.lower()
    return unit_mapping.get(normalized, unit_str)

def compare_values(base_value, converted_value, tolerance=0.1):
    """Compare two values with tolerance for numeric values"""
    # Handle None and empty values
    if base_value is None or base_value == "":
        base_value = ""
    if converted_value is None or converted_value == "":
        converted_value = ""
    
    # If both empty, consider as match
    if base_value == "" and converted_value == "":
        return True
    
    # Try numeric comparison first
    try:
        base_num = float(str(base_value).replace(',', '.'))
        conv_num = float(str(converted_value).replace(',', '.'))
        
        # Check if values are within tolerance (10% by default to account for rounding)
        if abs(base_num) < 0.0001 and abs(conv_num) < 0.0001:
            return True  # Both essentially zero
        elif abs(base_num) < 0.0001:
            return abs(conv_num) < tolerance
        else:
            relative_diff = abs(base_num - conv_num) / abs(base_num)
            return relative_diff < tolerance
    except (ValueError, TypeError):
        pass
    
    # String comparison - normalize and compare
    base_str = str(base_value).strip().lower()
    conv_str = str(converted_value).strip().lower()
    
    # Direct match
    if base_str == conv_str:
        return True
    
    # Check if one contains the other (for cases like "Yes" vs "True")
    if base_str in conv_str or conv_str in base_str:
        return True
    
    # Check for common equivalents
    equivalents = {
        'yes': ['true', 'yes', '1'],
        'no': ['false', 'no', '0'],
        'true': ['yes', 'true', '1'],
        'false': ['no', 'false', '0']
    }
    
    if base_str in equivalents:
        return conv_str in equivalents[base_str]
    if conv_str in equivalents:
        return base_str in equivalents[conv_str]
    
    return False

def load_base_dataset(file_path):
    """Load base dataset from Excel (format: Parameter | Data | Unit | Range)"""
    wb = openpyxl.load_workbook(file_path, data_only=True)  # data_only=True to get calculated values
    ws = wb.active
    
    base_data = {}
    
    # Read from row 2 (skip header)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # Skip empty rows
            continue
        
        parameter = row[0]  # Column A
        data = row[1] if len(row) > 1 else ""  # Column B
        unit = row[2] if len(row) > 2 else ""  # Column C
        range_val = row[3] if len(row) > 3 else ""  # Column D
        
        # Extract value and unit from data if unit embedded
        value, embedded_unit = extract_unit_from_value(str(data))
        if embedded_unit and not unit:
            unit = embedded_unit
        
        base_data[normalize_parameter_name(parameter)] = {
            'original_param': parameter,
            'value': value,
            'unit': normalize_unit(unit),
            'range': range_val
        }
    
    wb.close()
    return base_data

def load_converted_dataset(file_path):
    """Load converted dataset from Excel (format: Element ID | Element Type | Element Name | Parameter | Value | Data Type | Unit)"""
    wb = openpyxl.load_workbook(file_path, data_only=True)  # data_only=True to get calculated values
    ws = wb.active
    
    converted_data = {}
    
    # Read from row 2 (skip header)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[3]:  # Skip if no parameter name (column D)
            continue
        
        parameter = row[3]  # Column D - Parameter
        value = row[4] if len(row) > 4 else ""  # Column E - Value
        unit = row[6] if len(row) > 6 else ""  # Column G - Unit
        
        # Use normalized parameter as key
        norm_param = normalize_parameter_name(parameter)
        
        # Store only if not already stored (take first occurrence)
        if norm_param not in converted_data:
            converted_data[norm_param] = {
                'original_param': parameter,
                'value': value,
                'unit': normalize_unit(unit)
            }
    
    wb.close()
    return converted_data

def compare_datasets(base_data, converted_data):
    """Compare base and converted datasets"""
    comparison_results = []
    
    matched_count = 0
    total_count = len(base_data)
    
    for norm_param, base_info in base_data.items():
        result = {
            'parameter': base_info['original_param'],
            'base_value': base_info['value'],
            'base_unit': base_info['unit'],
            'converted_value': '',
            'converted_unit': '',
            'value_match': False,
            'unit_match': False,
            'found': False,
            'flag': 'Missing in Converted'
        }
        
        # Check if parameter exists in converted data
        if norm_param in converted_data:
            conv_info = converted_data[norm_param]
            result['found'] = True
            result['converted_value'] = conv_info['value']
            result['converted_unit'] = conv_info['unit']
            
            # Compare values
            result['value_match'] = compare_values(base_info['value'], conv_info['value'])
            
            # Compare units
            result['unit_match'] = (base_info['unit'] == conv_info['unit']) or \
                                  (not base_info['unit'] and not conv_info['unit'])
            
            # Determine flag
            if result['value_match'] and result['unit_match']:
                result['flag'] = 'Match'
                matched_count += 1
            elif result['value_match'] and not result['unit_match']:
                result['flag'] = 'Unit Mismatch'
            elif not result['value_match'] and result['unit_match']:
                result['flag'] = 'Value Mismatch'
            else:
                result['flag'] = 'Value & Unit Mismatch'
        
        comparison_results.append(result)
    
    # Calculate completeness
    completeness = (matched_count / total_count * 100) if total_count > 0 else 0
    
    return comparison_results, completeness, matched_count, total_count

def generate_excel_report(comparison_results, completeness, matched, total, output_file):
    """Generate Excel comparison report"""
    # Remove file if it exists to avoid conflicts
    if os.path.exists(output_file):
        try:
            os.remove(output_file)
        except:
            pass
    
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comparison Report"
    
    # Define colors
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Completeness colors
    if completeness >= 90:
        summary_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
    elif completeness >= 70:
        summary_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold
    else:
        summary_fill = PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid")  # Light salmon
    
    match_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
    no_match_fill = PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid")  # Light salmon
    missing_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold
    
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Title
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = "Dataset Comparison Report"
    title_cell.font = Font(bold=True, size=16, color="366092")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Summary
    ws.merge_cells('A2:H2')
    summary_cell = ws['A2']
    summary_cell.value = f"Completeness: {completeness:.1f}% ({matched}/{total} parameters matched)"
    summary_cell.font = Font(bold=True, size=12)
    summary_cell.fill = summary_fill
    summary_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 25
    
    # Headers
    headers = ['Parameter', 'Base Value', 'Base Unit', 'Converted Value', 'Converted Unit', 
               'Value Match', 'Unit Match', 'Status']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    ws.row_dimensions[4].height = 20
    
    # Data rows
    for row_num, result in enumerate(comparison_results, 5):
        ws.cell(row=row_num, column=1).value = sanitize_cell_value(result['parameter'])
        ws.cell(row=row_num, column=2).value = sanitize_cell_value(result['base_value'])
        ws.cell(row=row_num, column=3).value = sanitize_cell_value(result['base_unit'])
        ws.cell(row=row_num, column=4).value = sanitize_cell_value(result['converted_value'])
        ws.cell(row=row_num, column=5).value = sanitize_cell_value(result['converted_unit'])
        
        # Value Match column with color
        value_match_cell = ws.cell(row=row_num, column=6)
        value_match_cell.value = 'YES' if result['value_match'] else 'NO'
        value_match_cell.fill = match_fill if result['value_match'] else no_match_fill
        value_match_cell.font = Font(bold=True)
        value_match_cell.alignment = Alignment(horizontal='center')
        
        # Unit Match column with color
        unit_match_cell = ws.cell(row=row_num, column=7)
        unit_match_cell.value = 'YES' if result['unit_match'] else 'NO'
        unit_match_cell.fill = match_fill if result['unit_match'] else no_match_fill
        unit_match_cell.font = Font(bold=True)
        unit_match_cell.alignment = Alignment(horizontal='center')
        
        # Status column with color
        status_cell = ws.cell(row=row_num, column=8)
        status_cell.value = sanitize_cell_value(result['flag'])
        if result['flag'] == 'MATCH':
            status_cell.fill = match_fill
            status_cell.font = Font(bold=True, color="006400")  # Dark green
        elif result['flag'] == 'MISSING':
            status_cell.fill = missing_fill
            status_cell.font = Font(bold=True, color="8B4513")  # Saddle brown
        else:
            status_cell.fill = no_match_fill
            status_cell.font = Font(bold=True, color="8B0000")  # Dark red
        status_cell.alignment = Alignment(horizontal='center')
        
        # Apply borders to all cells in row
        for col in range(1, 9):
            ws.cell(row=row_num, column=col).border = border
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 13
    ws.column_dimensions['G'].width = 13
    ws.column_dimensions['H'].width = 20
    
    # Save workbook
    wb.save(output_file)
    print(f"Excel report generated: {output_file}")

def generate_pdf_report(comparison_results, completeness, matched, total, output_file):
    """Generate PDF comparison report"""
    doc = SimpleDocTemplate(output_file, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#366092'),
        spaceAfter=30,
        alignment=1  # Center
    )
    elements.append(Paragraph("Dataset Comparison Report", title_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Summary
    summary_style = ParagraphStyle(
        'Summary',
        parent=styles['Normal'],
        fontSize=14,
        textColor=colors.black,
        spaceAfter=20,
        alignment=1
    )
    
    # Color code summary based on completeness
    if completeness >= 90:
        summary_color = colors.HexColor('#C6EFCE')
    elif completeness >= 70:
        summary_color = colors.HexColor('#FFEB9C')
    else:
        summary_color = colors.HexColor('#FFC7CE')
    
    summary_text = f"<b>Completeness: {completeness:.1f}%</b><br/>({matched}/{total} parameters matched)"
    elements.append(Paragraph(summary_text, summary_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Prepare table data
    table_data = [['Parameter', 'Base Value', 'Base Unit', 'Conv. Value', 'Conv. Unit', 'Status']]
    
    for result in comparison_results:
        row = [
            result['parameter'][:30],  # Truncate long names
            str(result['base_value'])[:15],
            result['base_unit'],
            str(result['converted_value'])[:15],
            result['converted_unit'],
            result['flag']
        ]
        table_data.append(row)
    
    # Create table
    col_widths = [2.2*inch, 1*inch, 0.8*inch, 1*inch, 0.8*inch, 1.2*inch]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    # Table style
    table_style = TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Body
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ])
    
    # Color code rows based on status
    for idx, result in enumerate(comparison_results, 1):
        if result['flag'] == 'Match':
            table_style.add('BACKGROUND', (0, idx), (-1, idx), colors.HexColor('#C6EFCE'))
        elif result['flag'] == 'Missing in Converted':
            table_style.add('BACKGROUND', (0, idx), (-1, idx), colors.HexColor('#FFC7CE'))
        else:
            table_style.add('BACKGROUND', (0, idx), (-1, idx), colors.HexColor('#FFEB9C'))
    
    table.setStyle(table_style)
    elements.append(table)
    
    # Legend
    elements.append(Spacer(1, 0.3*inch))
    legend_style = ParagraphStyle(
        'Legend',
        parent=styles['Normal'],
        fontSize=9,
        spaceAfter=5
    )
    elements.append(Paragraph("<b>Legend:</b>", legend_style))
    elements.append(Paragraph("🟢 Green = Match | 🟡 Yellow = Mismatch | 🔴 Red = Missing", legend_style))
    
    # Build PDF
    doc.build(elements)
    print(f"PDF report generated: {output_file}")

def main():
    """Main execution function"""
    # Create root window and hide it
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    
    print("="*60)
    print("EXCEL DATASET COMPARISON TOOL")
    print("="*60)
    
    # Step 1: Select base dataset
    print("\nStep 1: Select the BASE dataset (DPP template format)")
    base_file = filedialog.askopenfilename(
        title="Select BASE Dataset Excel File",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        initialdir="."
    )
    
    if not base_file:
        print("No base file selected. Exiting...")
        messagebox.showwarning("No File Selected", "No base dataset file was selected. Operation cancelled.")
        return
    
    print(f"✓ Base dataset: {os.path.basename(base_file)}")
    
    # Step 2: Select converted dataset
    print("\nStep 2: Select the CONVERTED dataset (IFC conversion output)")
    converted_file = filedialog.askopenfilename(
        title="Select CONVERTED Dataset Excel File",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        initialdir=os.path.dirname(base_file)
    )
    
    if not converted_file:
        print("No converted file selected. Exiting...")
        messagebox.showwarning("No File Selected", "No converted dataset file was selected. Operation cancelled.")
        return
    
    print(f"✓ Converted dataset: {os.path.basename(converted_file)}")
    
    # Step 3: Ask for output name
    print("\nStep 3: Enter output file name")
    output_name = simpledialog.askstring(
        "Output File Name",
        "Enter a name for the comparison report (without extension):",
        parent=root,
        initialvalue="comparison_report"
    )
    
    if not output_name:
        output_name = "comparison_report"
    
    # Remove extensions if user added them
    output_name = output_name.replace('.xlsx', '').replace('.pdf', '')
    
    # Step 4: Select output directory
    print("\nStep 4: Select output folder")
    output_dir = filedialog.askdirectory(
        title="Select folder to save comparison reports",
        initialdir=os.path.dirname(base_file)
    )
    
    if not output_dir:
        print("No output directory selected. Using base file directory...")
        output_dir = os.path.dirname(base_file)
    
    print(f"✓ Output directory: {output_dir}")
    
    # Create output file paths
    excel_output = os.path.join(output_dir, f"{output_name}.xlsx")
    pdf_output = os.path.join(output_dir, f"{output_name}.pdf")
    
    print("\n" + "="*60)
    print("PROCESSING...")
    print("="*60)
    
    try:
        # Load datasets
        print("\n[1/5] Loading base dataset...")
        base_data = load_base_dataset(base_file)
        print(f"  ✓ Loaded {len(base_data)} parameters from base dataset")
        
        print("\n[2/5] Loading converted dataset...")
        converted_data = load_converted_dataset(converted_file)
        print(f"  ✓ Loaded {len(converted_data)} parameters from converted dataset")
        
        # Compare datasets
        print("\n[3/5] Comparing datasets...")
        comparison_results, completeness, matched, total = compare_datasets(base_data, converted_data)
        print(f"  ✓ Comparison complete: {completeness:.1f}% match rate")
        
        # Generate Excel report
        print("\n[4/5] Generating Excel report...")
        generate_excel_report(comparison_results, completeness, matched, total, excel_output)
        print(f"  ✓ Excel report saved")
        
        # Generate PDF report
        print("\n[5/5] Generating PDF report...")
        generate_pdf_report(comparison_results, completeness, matched, total, pdf_output)
        print(f"  ✓ PDF report saved")
        
        # Summary
        print("\n" + "="*60)
        print("COMPARISON COMPLETE!")
        print("="*60)
        print(f"\nResults:")
        print(f"  Total parameters: {total}")
        print(f"  Matched: {matched}")
        print(f"  Completeness: {completeness:.1f}%")
        print(f"\nReports saved:")
        print(f"  Excel: {excel_output}")
        print(f"  PDF: {pdf_output}")
        print("="*60)
        
        # Show success message
        messagebox.showinfo(
            "Comparison Complete",
            f"Dataset comparison successful!\n\n"
            f"Completeness: {completeness:.1f}%\n"
            f"Matched: {matched}/{total} parameters\n\n"
            f"Reports saved:\n"
            f"• {os.path.basename(excel_output)}\n"
            f"• {os.path.basename(pdf_output)}\n\n"
            f"Location:\n{output_dir}"
        )
        
    except Exception as e:
        print(f"\nERROR during comparison: {str(e)}")
        import traceback
        traceback.print_exc()
        messagebox.showerror("Comparison Error", f"An error occurred during comparison:\n\n{str(e)}")
    
    root.destroy()

if __name__ == '__main__':
    main()
